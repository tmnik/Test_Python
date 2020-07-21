import smtplib
import os
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from platform import python_version
from openpyxl import load_workbook


def count_lines():
    wb = load_workbook("test.xlsx")
    sheets_list = wb.sheetnames
    sheet_active = wb[sheets_list[0]]
    count = sheet_active.max_row
    return count


def message():
    n = count_lines()
    if 5 <= n <= 20 or n % 10 == 0:
        return f'В таблице Excel {str(n)} строк.'
    elif n % 10 == 1:
        return f'В таблице Excel {str(n)} строка.'
    elif 2 <= n % 10 <= 4:
        return f'В таблице Excel {str(n)} строки.'
    else:
        return f'В таблице Excel {str(n)} строк.'


server = 'smtp.mail.ru'
user = 'testpython1@mail.ru'
password = '1eS1Pu1#0n'

recipients = ['nikonenkotm@mail.ru']
sender = 'testpython1@mail.ru'
subject = 'Test. Report'

text = message()
html = '<html><head></head><body><p>' + text + '</p></body></html>'

filepath = "Report_NikonenkoTM.pdf"
basename = os.path.basename(filepath)
filesize = os.path.getsize(filepath)

msg = MIMEMultipart('alternative')
msg['Subject'] = subject
msg['From'] = 'Nikonenko Tatiana <' + sender + '>'
msg['To'] = ', '.join(recipients)
msg['Reply-To'] = sender
msg['Return-Path'] = sender
msg['X-Mailer'] = 'Python/' + (python_version())

part_text = MIMEText(text, 'plain')
part_html = MIMEText(html, 'html')
part_file = MIMEBase('application', 'octet-stream; name="{}"'.format(basename))
part_file.set_payload(open(filepath, "rb").read())
part_file.add_header('Content-Description', basename)
part_file.add_header('Content-Disposition', 'attachment; filename="{}"; size={}'.format(basename, filesize))
encoders.encode_base64(part_file)

msg.attach(part_text)
msg.attach(part_html)
msg.attach(part_file)

mail = smtplib.SMTP_SSL(server)
mail.login(user, password)
mail.sendmail(sender, recipients, msg.as_string())
mail.quit()
